import streamlit as st
import datetime
from dateutil.relativedelta import relativedelta
import json
import re
from google import genai
from google.genai import types
import openpyxl
import pypdf
from io import BytesIO

# --- ページ設定 ---
st.set_page_config(page_title="接続検討申込 自動転記システム", page_icon="🚀", layout="wide")

# --- 日付計算関数 ---
def calculate_dates():
    today = datetime.date.today()
    half_year = today + relativedelta(months=6)
    if half_year.month == 12:
        next_month_first = datetime.date(half_year.year + 1, 1, 1)
    else:
        next_month_first = datetime.date(half_year.year, half_year.month + 1, 1)
    last_day = next_month_first - datetime.timedelta(days=1)
    return today.strftime("%Y/%m/%d"), last_day.strftime("%Y/%m/%d")

# --- メインUI ---
st.title("🚀 接続検討申込 自動転記システム")
st.markdown("設定項目を入力し、必要なファイルをアップロードしてから「実行」ボタンを押してください。")

# サイドバー: APIキーとファイルアップロード
with st.sidebar:
    st.header("🔑 初期設定")
    api_key = st.text_input("Gemini APIキー", type="password", help="AIを利用するためのAPIキーを入力してください。")
    
    st.header("📁 ファイルアップロード")
    template_file = st.file_uploader("エクセルのテンプレートをアップロード", type=["xlsx"])
    pdf_files = st.file_uploader("謄本（全部事項証明書）のPDFをアップロード", type=["pdf"], accept_multiple_files=True)

# メインエリア: 設定項目
col1, col2 = st.columns(2)

with col1:
    st.subheader("📂 1. 申請の種類")
    app_type = st.radio("申請の種類を選ぶ", ["1: 蓄電池（系統用）", "2: 太陽光（高圧・特別高圧）"])
    is_battery = "蓄電池" in app_type

    st.subheader("🔋 3. 蓄電池の設定（蓄電池のみ）")
    battery_capacity = st.number_input("PCS1台あたりの蓄電池容量 (kWh)", min_value=0, value=514, disabled=not is_battery)

with col2:
    st.subheader("⚙️ 2. 機器の設定（共通）")
    pcs_type = st.selectbox("PCSの種類", ["1: Sungrow SC125CX", "2: Huawei SUN2000-50KTL-NHM3"])
    total_pcs = st.number_input("全体のPCS設置台数", min_value=1, value=16)
    logic_choice = st.radio("特高回避ロジック", ["1: 適用する（1999kW制限・メイン機+制限機1台）", "2: 適用しない（制限なし）"])

    st.subheader("☀️ 4. 太陽光の設定（太陽光のみ）")
    solar_output = st.number_input("太陽光パネルの合計出力 (kW)", min_value=0, value=2500, disabled=is_battery)

# --- 実行ボタン ---
if st.button("▶ 処理を実行する", type="primary", use_container_width=True):
    if not api_key:
        st.error("🚨 APIキーが入力されていません。")
        st.stop()
    if not template_file:
        st.error("🚨 エクセルのテンプレートがアップロードされていません。")
        st.stop()
    if not pdf_files:
        st.error("🚨 PDFファイルがアップロードされていません。")
        st.stop()

    try:
        client = genai.Client(api_key=api_key)
    except Exception as e:
        st.error("🚨 APIキーの認証に失敗しました。正しいキーか確認してください。")
        st.stop()

    # 処理開始
    with st.spinner("処理を実行中です...しばらくお待ちください。"):
        selected_pcs_key = pcs_type.split(":")[0]
        logic_id = logic_choice.split(":")[0]
        
        PCS_DATABASE = {
            "1": {"maker": "Sungrow Power Supply Co., Ltd.", "model": "SC125CX", "discharge": 125, "charge": 122},
            "2": {"maker": "Huawei Technologies Co., Ltd.", "model": "SUN2000-50KTL-NHM3", "discharge": 50, "charge": 50}
        }
        pcs_data = PCS_DATABASE[selected_pcs_key]

        # エクセル読み込み
        wb = openpyxl.load_workbook(template_file, keep_vba=True)
        ws_hajimeni = next((wb[s] for s in wb.sheetnames if "はじめに" in s), None)
        ws_nyuryoku = next((wb[s] for s in wb.sheetnames if "入力" in s), None)
        ws_yashiki6 = next((wb[s] for s in wb.sheetnames if "様式６" in s), None)

        today_str, target_date_str = calculate_dates()

        # --- エクセル書き込みロジック（基本設定） ---
        if ws_hajimeni:
            ws_hajimeni["AV39"] = "はい"
            ws_hajimeni["AV42"] = "確認のうえ記載する"
            ws_hajimeni["AV71"] = "確認のうえ記載する"

        if ws_nyuryoku:
            ws_nyuryoku["E10"] = today_str
            ws_nyuryoku["E11"] = "未定"
            ws_nyuryoku["E12"] = "270-0091"
            ws_nyuryoku["E13"] = "千葉県"
            ws_nyuryoku["E14"] = "松戸市本町11-5 明治安田生命ビル1階"
            ws_nyuryoku["E15"] = "エイタイジャパンカブシキガイシャ ダイヒョウトリシマリヤク キクブングン"
            ws_nyuryoku["E16"] = "盈泰ジャパン株式会社"
            ws_nyuryoku["E17"] = "代表取締役 鞠 文軍"
            ws_nyuryoku["E18"] = "エイタイジャパンカブシキガイシャ"
            ws_nyuryoku["E19"] = "盈泰ジャパン株式会社"
            ws_nyuryoku["E20"] = "無"
            ws_nyuryoku["E25"] = "無"
            ws_nyuryoku["E26"] = "全量売電"
            ws_nyuryoku["E29"] = "新規"
            ws_nyuryoku["E32"] = int(60000000)
            ws_nyuryoku["E34"] = "相違なし"
            ws_nyuryoku["E35"] = "有"

            ws_nyuryoku["E36"] = "271-0092"
            ws_nyuryoku["E37"] = "千葉県"
            ws_nyuryoku["E38"] = "松戸市松戸1307-1 松戸ビル11階"
            ws_nyuryoku["E39"] = "盈泰ジャパン株式会社"
            ws_nyuryoku["E40"] = "開発部"
            ws_nyuryoku["E52"] = "佐野 華子"

            ws_nyuryoku["E60"] = target_date_str
            ws_nyuryoku["E61"] = target_date_str
            ws_nyuryoku["E62"] = target_date_str
            ws_nyuryoku["E63"] = "無"
            ws_nyuryoku["E67"] = "有（その他負荷「無」）"
            ws_nyuryoku["E68"] = "希望する"
            ws_nyuryoku["E80"] = "✓"
            ws_nyuryoku["E81"] = "【連絡先】の記載と同じ"
            
            ws_nyuryoku["E194"] = int(3 * total_pcs)
            ws_nyuryoku["E195"] = int(1)
            ws_nyuryoku["E191"] = "無"

            if is_battery:
                ws_nyuryoku["E90"] = "新設"
                ws_nyuryoku["E88"] = int(pcs_data["discharge"])
                ws_nyuryoku["E89"] = int(pcs_data["charge"])
                ws_nyuryoku["E92"] = pcs_data["maker"]
                ws_nyuryoku["E93"] = pcs_data["model"]
                ws_nyuryoku["E94"] = "三相３線式"
                ws_nyuryoku["E105"] = "自励式（電圧形）"
                ws_nyuryoku["E106"] = "有"
                ws_nyuryoku["E107"] = "無"
                ws_nyuryoku["E111"] = int(1)
                ws_nyuryoku["E112"] = "【PCS1台あたりの蓄電池容量】"

                max_consumption = 3 * total_pcs
                if logic_id == "1":
                    set_count_main = total_pcs - 1
                    limited_discharge = 1999 - (pcs_data["discharge"] * set_count_main)
                    limited_charge = 1999 - (pcs_data["charge"] * set_count_main) - max_consumption
                    
                    ws_nyuryoku["E113"] = int(battery_capacity)
                    ws_nyuryoku["H113"] = int(set_count_main)
                    ws_nyuryoku["H118"] = int(set_count_main)
                    ws_nyuryoku["E120"] = "有"
                    ws_nyuryoku["H121"] = int(limited_discharge)
                    ws_nyuryoku["H122"] = int(limited_charge)
                    ws_nyuryoku["E123"] = "新設"
                    ws_nyuryoku["E125"] = pcs_data["maker"]
                    ws_nyuryoku["E126"] = pcs_data["model"]
                    ws_nyuryoku["E146"] = int(battery_capacity)
                    ws_nyuryoku["H146"] = int(1)
                else:
                    ws_nyuryoku["E113"] = int(battery_capacity)
                    ws_nyuryoku["H113"] = int(total_pcs)
                    ws_nyuryoku["H118"] = int(total_pcs)
                    ws_nyuryoku["E120"] = "無"
            else:
                ws_nyuryoku["E90"] = "新設"

        # --- AI解析と様式6への転記 ---
        EXTRACTION_PROMPT = """
        あなたは日本の電力申請における土地調査の専門アシスタントです。
        提示された登記簿PDFのテキストデータから、以下のルールを「1文字の例外もなく」遵守して正確な情報を抽出し、JSON形式で出力してください。

        【絶対遵守の抽出ガードレール】
        1. 住所（address）: 登記簿の「所在」と「地番」を正確に結合すること。必ず先頭に正しい「都道府県名」を付与すること。
        2. 地目（chaimoku）: 登記簿の表題部にある地目をそのまま抽出すること。
        3. 所有者（owner）: 甲区にある現在の「所有者」の氏名。※最新の順位番号から抽出。共有の場合は全員の氏名を改行(\\n)で繋ぐ。

        【出力JSONフォーマット】
        {
          "address": "",
          "chaimoku": "",
          "owner": ""
        }
        ---
        【テキスト】
        __TEXT__
        """

        row_idx = 15
        last_address = "【住所】"
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for idx, pdf in enumerate(pdf_files):
            status_text.text(f"PDF解析中 ({idx + 1}/{len(pdf_files)}): {pdf.name}")
            extracted_text = ""
            try:
                pdf_reader = pypdf.PdfReader(pdf)
                for page in pdf_reader.pages:
                    extracted_text += page.extract_text() or ""
            except Exception as e:
                st.warning(f"{pdf.name} の読み込みに失敗しました。スキップします。")
                continue

            prompt_filled = EXTRACTION_PROMPT.replace("__TEXT__", extracted_text[:8000])
            try:
                extract_res = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=prompt_filled,
                    config=types.GenerateContentConfig(response_mime_type="application/json"),
                )
                data = json.loads(extract_res.text)
            except Exception as e:
                st.warning(f"{pdf.name} のAI解析に失敗しました。スキップします。")
                continue

            if ws_yashiki6:
                ws_yashiki6.cell(row=row_idx, column=14, value=int(idx + 1))
                ws_yashiki6.cell(row=row_idx, column=17, value=data.get("address", ""))
                ws_yashiki6.cell(row=row_idx, column=35, value="その他")
                ws_yashiki6.cell(row=row_idx, column=40, value="民地")
                ws_yashiki6.cell(row=row_idx, column=44, value=data.get("chaimoku", ""))
                ws_yashiki6.cell(row=row_idx, column=48, value=data.get("owner", ""))

            last_address = data.get('address', last_address)
            row_idx += 1
            progress_bar.progress((idx + 1) / len(pdf_files))

        # 備忘録の書き込み
        if ws_yashiki6:
            if is_battery:
                reminder_path = f"G:\\共有ドライブ\\蓄電所_01.高圧\\{last_address}（蓄電所）\\10_土地関係(または関連)\\謄本・公図"
            else:
                reminder_path = f"G:\\共有ドライブ\\太陽光_01.高圧\\{last_address}（太陽光）\\10_土地関係(または関連)\\謄本・公図"
            ws_yashiki6["A1"] = f"📁 【備忘録】提出用PDFの格納先: {reminder_path}"

        # メモリ上にエクセルを保存
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        status_text.text("✅ 処理が完了しました！")
        st.success("エクセルの生成が完了しました！下のボタンからダウンロードしてください。")
        
        output_filename = "転記完了_蓄電池_接続検討申込.xlsx" if is_battery else "転記完了_太陽光_接続検討申込.xlsx"
        
        st.download_button(
            label="📥 完成したエクセルをダウンロード",
            data=output_buffer,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )